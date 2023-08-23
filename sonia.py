import datetime
from dateutil import parser
import openpyxl as xl
import pandas as pd
import requests
from openpyxl.styles import PatternFill

cur_date = '22.08.23'
cur_month = 'AUG'
cur_year = 2023

cur_date_datetime = parser.parse(cur_date).date()

key = "bf5204b93cd4e38625e4d899fc6d5e9f"

# shares = ['AAPL', 'AMZN', 'META', 'MSFT', 'NFLX', 'NVDA', 'QQQ', 'TSLA']

shares = ['AAPL']

for share in shares:
    url = rf'https://financialmodelingprep.com/api/v3/historical-chart/1min/{share}?apikey={key}'
    path = rf'E:\sonia daily data\1 min cash\{cur_year}\{cur_month}\{cur_date}\{share} 1 min csh.xlsx'

    # response = requests.get(url)
    # data = response.json()

    data = pd.read_json(r'C:\Users\admin\PycharmProjects\daily data\AAPL.json')

    df = pd.DataFrame(data)

    # df["date"] = pd.to_datetime(df["date"])
    # df["Date"] = df["date"].dt.date
    # df["Time"] = df["date"].dt.time
    #
    # df.drop(df[(df.Date < cur_date_datetime)].index, inplace=True)
    # df.drop(df[(df.Date > cur_date_datetime)].index, inplace=True)
    #
    # df = df.iloc[:, [7, 3, 2, 4, 5]]
    # df = df.sort_values(by='Time')
    #
    # df = df.round(2)

    with pd.ExcelWriter(path) as writer:
        df.to_excel(writer, index=False)

    df.to_json(rf'C:\Users\admin\PycharmProjects\daily data\AAPL.json')

    # excel formatting
    wb = xl.load_workbook(path)
    sheet = wb['Sheet1']

    start_row = 2

    high_dict = {'AAPL': 0, 'AMZN': 0, 'META': 0, 'MSFT': 0, 'NFLX': 0, 'NVDA': 0, 'QQQ': 0, 'TSLA:': 0}
    low_dict = {'AAPL': 0, 'AMZN': 0, 'META': 0, 'MSFT': 0, 'NFLX': 0, 'NVDA': 0, 'QQQ': 0, 'TSLA:': 0}
    cl_9_40_dict = {'AAPL': 0, 'AMZN': 0, 'META': 0, 'MSFT': 0, 'NFLX': 0, 'NVDA': 0, 'QQQ': 0, 'TSLA:': 0}

    # converting time from str to datetime.datetime
    while start_row <= len(sheet['A']):
        time_cell = sheet.cell(start_row, 1)

        time = time_cell.value

        time = datetime.datetime.strptime(time, "%H:%M:%S")
        # time = time.time()
        time_cell.value = time
        time_cell.number_format = 'h:mm AM/PM'

        start_row += 1

    wb.save(path)

    # reloading wb
    wb = xl.load_workbook(path)
    sheet = wb['Sheet1']

    start_row = 2
    time_cell = sheet.cell(start_row, 1)
    cur_time = time_cell.value
    # print(cur_time)

    while time_cell.value < datetime.datetime(1900, 1, 1, hour=9, minute=40):
        start_row += 1
        time_cell = sheet.cell(start_row, 1)

    print(f"starting row is {start_row}")
    start_row_2 = start_row

    time_cell = sheet.cell(start_row, 1)
    cur_time = time_cell.value
    end_time = datetime.datetime(1900, 1, 1,16, 0, 0)

    # 9:40 close value
    cl_9_40_dict[share] = sheet.cell(start_row, 4).value

    # 9:40 AM row formatting
    for i in range(1, 6):
        sheet.cell(start_row, i).fill = PatternFill("solid", 'FFFF00')

    # reloading wb otherwise pattern fill doesn't work
    wb.save(path)
    wb = xl.load_workbook(path)
    sheet = wb['Sheet1']

    HIGH = 0
    LOW = 9999999

    # HIGH and LOW value finding loop
    while cur_time is not None and cur_time <= end_time:
        time_cell = sheet.cell(start_row, 1)
        high_cell = sheet.cell(start_row, 2)
        low_cell = sheet.cell(start_row, 3)

        cur_time = time_cell.value

        if high_cell.value is not None and high_cell.value > HIGH:
            HIGH = high_cell.value

        if low_cell.value is not None and low_cell.value < LOW and low_cell.value != 0:
            LOW = low_cell.value

        start_row += 1

    high_dict[share] = HIGH
    low_dict[share] = LOW

    # 30 MIN FORMATTING IN 1 MIN SHEETS
    HIGH = 0
    LOW = 9999999

    sheet.cell(1, 7).value = "HIGH"
    sheet.cell(1, 8).value = "LOW"
    sheet.cell(1, 9).value = "CLOSE"

    start_row = start_row_2  # actual start row

    time_cell = sheet.cell(start_row, 1)
    cur_time = time_cell.value

    count = 0

    while cur_time is not None and cur_time <= end_time:
        high_cell = sheet.cell(start_row, 2)
        low_cell = sheet.cell(start_row, 3)

        # print(cur_time)

        if high_cell.value is not None and high_cell.value > HIGH:
            HIGH = high_cell.value

        if low_cell.value is not None and low_cell.value < LOW and low_cell.value != 0:
            LOW = low_cell.value

        # resetting after 30 mins
        if count == 30:
            sheet.cell(start_row, 7).value = HIGH
            sheet.cell(start_row, 8).value = LOW

            # if 30 min close is empty or 0
            if sheet.cell(start_row, 4).value == 0 or sheet.cell(start_row, 4).value is None:
                temp_row = start_row

                while sheet.cell(temp_row, 4).value == 0 or sheet.cell(temp_row, 4).value is None:
                    temp_row -= 1

                sheet.cell(start_row, 9).value = sheet.cell(temp_row, 4).value  # close

            else:
                sheet.cell(start_row, 9).value = sheet.cell(start_row, 4).value  # close

            count = 1
            HIGH = 0
            LOW = 9999999
            start_row += 1
            continue

        start_row += 1
        count += 1

        time_cell = sheet.cell(start_row, 1)
        cur_time = time_cell.value

    # last any left aggregate (< 30 mins)
    sheet.cell(start_row - 1, 7).value = HIGH
    sheet.cell(start_row - 1, 8).value = LOW
    sheet.cell(start_row - 1, 9).value = sheet.cell(start_row - 1, 4).value  # close

    print(f"{share} done")

    # converting datetime.datetime to str in specific format (%I:%M %p)
    start_row = 2

    while start_row < len(sheet['A']):
        time_cell = sheet.cell(start_row, 1)
        time = time_cell.value
        time_cell.value = time.strftime("%I:%M %p")
        time_cell.number_format = 'h:mm AM/PM'

        start_row += 1

    wb.save(path)

    # print(high_dict, low_dict, cl_9_40_dict)





# 30 min
# for share in shares:
#     url = rf'https://financialmodelingprep.com/api/v3/historical-chart/30min/{share}?apikey={key}'
#     path = rf'E:\sonia daily data\30 min cash\{cur_year}\{cur_month}\{cur_date}\{share} 30 min csh.xlsx'
#
#     response = requests.get(url)
#
#     data = response.json()
#
#     df = pd.DataFrame(data)
#     df["date"] = pd.to_datetime(df["date"])
#     df["Date"] = df["date"].dt.date
#     df["Time"] = df["date"].dt.time
#
#     df.drop(df[df.Date < cur_date_datetime].index, inplace=True)
#
#     df = df.iloc[:, [7, 1, 3, 2, 4, 5]]
#     df = df.sort_values(by='Time')
#
#     df.to_excel(path, index=False)
#