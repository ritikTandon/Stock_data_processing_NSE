import openpyxl as xl
from openpyxl.styles import Font, Alignment
import datetime
from date_variables import date, mnth, yr

cashHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\cash high low.xlsx')
cashHL_sheet = cashHL_wb['Sheet1']

foHL_wb = xl.load_workbook(r'C:\Users\admin\PycharmProjects\daily data\fo high low.xlsx')
foHL_sheet = foHL_wb['Sheet1']

# cash_30_min_list = ["ADANI", "APOLLO", "BAJFINSV", "BAJFIN", "BANBK", "BARODA", "BN", "DLF", "EICHER", "FEDBANK", "HCL",
#                     "INDUSIND", "JIND", "LIC", "M&M", "M&MFIN", "NIFTY", "SBIN", "SUNTV", "TM", "TP", "TS"]

# index for getting values from cash/fo high low sheets
index_30_min = [2, 3, 4, 5, 6, 7, 4, 9, 10, 11, 12, 16, 18, 19, 20, 21, 10, 23, 24, 26, 27, 28]

cash_30_min_list = ["ADANI"]

idx = 0

# LTP and PREV
ltp_wb = xl.load_workbook(rf'C:\Users\admin\PycharmProjects\daily data\LTP PREV.xlsx')
ltp_sheet = ltp_wb["ltp"]
ltp_row = 2

while ltp_row < len(ltp_sheet["A"]):
    ltp_sheet.cell(ltp_row, 3).value = ltp_sheet.cell(ltp_row, 2).value  # moving last day's LTP to 'PREV'

    if ltp_sheet.cell(ltp_row, 1).value in ["BN", "NIFTY"]:    # separate for NIFTY and BN as their data LTP come from 'fo high low.xlsx'
        ltp_sheet.cell(ltp_row, 2).value = foHL_sheet.cell(index_30_min[idx], 5).value

    else:
        ltp_sheet.cell(ltp_row, 2).value = cashHL_sheet.cell(index_30_min[idx], 5).value

    ltp_row += 1
    idx += 1

idx = 0

ltp_wb.save(rf'C:\Users\admin\PycharmProjects\daily data\LTP PREV.xlsx')


for share in cash_30_min_list:
    path = rf"E:\Daily Data work\hourlys 30 minute CASH\{yr}\{mnth}\{date}\{share}.xlsx"

    cash_30_min_wb = xl.load_workbook(path)
    old_30_min_sheet = cash_30_min_wb[f"{share}-Sheet1"]

    new_30_min_sheet = cash_30_min_wb.create_sheet(f"{share}")

    # FIXED HEADINGS
    new_30_min_sheet.cell(6, 6).value = f'{share}'
    new_30_min_sheet.cell(6, 7).value = "HIGH"
    new_30_min_sheet.cell(6, 8).value = "LOW"
    new_30_min_sheet.cell(6, 9).value = "LTP"
    new_30_min_sheet.cell(6, 10).value = "PREV"

    new_30_min_sheet.cell(8, 6).value = "Time"
    new_30_min_sheet.cell(8, 7).value = "High Rate"
    new_30_min_sheet.cell(8, 8).value = "Low Rate"
    new_30_min_sheet.cell(8, 9).value = "Close Rate"

    old_sheet_row = 2

    while old_sheet_row <= len(old_30_min_sheet["A"]):
        # time
        new_30_min_sheet.cell(old_sheet_row + 7, 6).value = old_30_min_sheet.cell(old_sheet_row, 7).value
        new_30_min_sheet.cell(old_sheet_row + 7, 6).number_format = 'hh:mm AM/PM'

        # high
        new_30_min_sheet.cell(old_sheet_row + 7, 7).value = old_30_min_sheet.cell(old_sheet_row, 4).value

        # low
        new_30_min_sheet.cell(old_sheet_row + 7, 8).value = old_30_min_sheet.cell(old_sheet_row, 5).value

        # close
        new_30_min_sheet.cell(old_sheet_row + 7, 9).value = old_30_min_sheet.cell(old_sheet_row, 3).value

        old_sheet_row += 1

    # bolding the sheet
    for i in range(25):
        for j in range(15):
            new_30_min_sheet.cell(i+1, j+1).font = Font(bold=True)
            new_30_min_sheet.cell(i+1, j+1).alignment = Alignment(horizontal='center')

    # Filling the data
    if share in ["BN", "NIFTY"]:    # separate for NIFTY and BN as their data will come from 'fo high low.xlsx'
        new_30_min_sheet.cell(7, 6).value = cashHL_sheet.cell(index_30_min[idx], 7).value   # 9:25 cl
        new_30_min_sheet.cell(7, 7).value = cashHL_sheet.cell(index_30_min[idx], 2).value   # HIGH
        new_30_min_sheet.cell(7, 8).value = cashHL_sheet.cell(index_30_min[idx], 3).value   # LOW
        # new_30_min_sheet.cell(7, 9).value = cashHL_sheet.cell(index_30_min[idx], 7).value   # LTP
        # new_30_min_sheet.cell(7, 10).value = cashHL_sheet.cell(index_30_min[idx], 7).value   # PREV

    else:
        pass

    idx += 1

    cash_30_min_wb.save(rf'C:\Users\admin\PycharmProjects\daily data\{share}1.xlsx')
